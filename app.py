import io
import os
import smtplib
from email.message import EmailMessage

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill

st.set_page_config(page_title="화장품 영업 데이터 분석", page_icon="💄", layout="wide")
st.title("💄 화장품 영업 데이터 분석")
st.caption("엑셀을 업로드하면 요약 시트와 그래프가 포함된 가공 파일을 생성합니다.")
CHATBOT_MODEL = "gpt-4o-mini"

def secret(name, default=""):
    try:
        return st.secrets.get(name, default)
    except Exception:
        return default

def find_col(df, names):
    for c in df.columns:
        s = str(c).replace(" ", "").lower()
        if any(n.replace(" ", "").lower() in s for n in names): return c
    return None

def build_workbook(upload):
    raw = pd.read_excel(upload, sheet_name=0)
    date_col = find_col(raw, ["날짜", "일자", "date"])
    product_col = find_col(raw, ["제품명", "상품명", "product"])
    sales_col = find_col(raw, ["매출액", "매출", "sales", "revenue"])
    profit_rate_col = find_col(raw, ["영업이익률", "이익률", "profitmargin", "margin"])
    missing = [label for label, col in [("날짜", date_col), ("제품명", product_col), ("매출액", sales_col), ("영업이익률", profit_rate_col)] if col is None]
    if missing: raise ValueError("필수 열을 찾을 수 없습니다: " + ", ".join(missing))
    raw[date_col] = pd.to_datetime(raw[date_col], errors="coerce").dt.date
    raw[sales_col] = pd.to_numeric(raw[sales_col].astype(str).str.replace(",", ""), errors="coerce").fillna(0)
    raw[profit_rate_col] = pd.to_numeric(raw[profit_rate_col].astype(str).str.replace("%", ""), errors="coerce")
    by_date = raw.groupby(date_col, dropna=False).agg(매출액=(sales_col, "sum"), 평균매출액=(sales_col, "mean"), 평균영업이익률=(profit_rate_col, "mean")).reset_index()
    by_product = raw.groupby(product_col, dropna=False).agg(매출액=(sales_col, "sum"), 평균영업이익률=(profit_rate_col, "mean")).reset_index()
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        raw.to_excel(writer, index=False, sheet_name="원본데이터")
        by_date.to_excel(writer, index=False, sheet_name="날짜별매출액")
        by_product.to_excel(writer, index=False, sheet_name="제품명별매출액")
    out.seek(0); wb = load_workbook(out)
    ws = wb.create_sheet("그래프")
    ws["A1"] = "제품명별 평균 영업이익률"; ws["A1"].font = Font(bold=True, size=14)
    for i, row in enumerate(by_product.itertuples(index=False), 2): ws.cell(i, 1, row[0]); ws.cell(i, 2, row[2])
    chart = BarChart(); chart.title = "제품명 별 평균 영업이익률"; chart.y_axis.title = "영업이익률 (%)"; chart.x_axis.title = "제품명"
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=len(by_product)+1), titles_from_data=True); chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=len(by_product)+1)); chart.anchor = "G5"; ws.add_chart(chart)
    start = len(by_product) + 5; ws.cell(start, 1, "날짜별 평균 매출액").font = Font(bold=True, size=14)
    for j, row in enumerate(by_date.itertuples(index=False), start + 1): ws.cell(j, 1, row[0]); ws.cell(j, 2, row[2])
    line = LineChart(); line.title = "날짜 별 평균 매출액"; line.y_axis.title = "평균 매출액"; line.x_axis.title = "날짜"; line.add_data(Reference(ws, min_col=2, min_row=start, max_row=start+len(by_date)), titles_from_data=True); line.set_categories(Reference(ws, min_col=1, min_row=start+1, max_row=start+len(by_date))); line.anchor = "G10"; ws.add_chart(line)
    final = io.BytesIO(); wb.save(final); final.seek(0)
    return raw, by_date, by_product, final.getvalue()

upload = st.file_uploader("화장품_영업_데이터 엑셀 파일 업로드", type=["xlsx", "xls"])
if upload:
    try:
        raw, by_date, by_product, processed = build_workbook(upload)
        st.success(f"처리 완료: {len(raw):,}건")
        c1, c2 = st.columns(2); c1.dataframe(by_date, use_container_width=True); c2.dataframe(by_product, use_container_width=True)
        st.download_button("⬇️ 가공된 엑셀 다운로드", processed, "화장품_영업_데이터_가공본.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.session_state["processed"] = processed
        st.session_state["context"] = (raw.head(100).to_csv(index=False), by_date.to_csv(index=False), by_product.to_csv(index=False))
    except Exception as e: st.error(str(e))

tab1, tab2 = st.tabs(["📧 엑셀 메일 발송", "💬 데이터 챗봇"])
with tab1:
    if "processed" not in st.session_state: st.info("먼저 엑셀 파일을 업로드하세요.")
    else:
        to = st.text_input("받는 사람 이메일"); subject = st.text_input("제목", "화장품 영업 데이터 가공본"); body = st.text_area("본문", "가공된 엑셀 파일을 첨부드립니다.")
        if st.button("메일 발송"):
            try:
                msg = EmailMessage(); msg["Subject"] = subject; msg["From"] = secret("SMTP_USER"); msg["To"] = to; msg.set_content(body); msg.add_attachment(st.session_state["processed"], maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="화장품_영업_데이터_가공본.xlsx")
                with smtplib.SMTP(secret("SMTP_HOST", "smtp.gmail.com"), int(secret("SMTP_PORT", "587"))) as smtp: smtp.starttls(); smtp.login(secret("SMTP_USER"), secret("SMTP_PASSWORD")); smtp.send_message(msg)
                st.success("메일을 발송했습니다.")
            except Exception as e: st.error(f"발송 실패: {e}. SMTP_USER/SMTP_PASSWORD 등 설정을 확인하세요.")
with tab2:
    if "context" not in st.session_state: st.info("먼저 엑셀 파일을 업로드하세요.")
    else:
        if "messages" not in st.session_state: st.session_state.messages = []
        for m in st.session_state.messages: st.chat_message(m["role"]).write(m["content"])
        if q := st.chat_input("예: 매출액이 가장 큰 제품은 무엇인가요?"):
            st.session_state.messages.append({"role":"user", "content":q}); st.chat_message("user").write(q)
            key = secret("OPENAI_API_KEY")
            if key:
                from openai import OpenAI
                prompt = "다음 CSV 자료만 근거로 한국어로 답하세요.\n" + "\n".join(st.session_state["context"]) + "\n질문: " + q
                ans = OpenAI(api_key=key).chat.completions.create(model=CHATBOT_MODEL, messages=[{"role":"user","content":prompt}]).choices[0].message.content
            else: ans = "OPENAI_API_KEY가 설정되지 않아 챗봇을 사용할 수 없습니다. .streamlit/secrets.toml에 키를 설정하세요."
            st.session_state.messages.append({"role":"assistant", "content":ans}); st.chat_message("assistant").write(ans)
